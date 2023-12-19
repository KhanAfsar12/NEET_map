from flask import Flask, jsonify, redirect, request, url_for
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://postgres:root@localhost:5432/account'
db = SQLAlchemy(app)


# data_subject = db.Table('data_subject',
#                         db.Column('college_code', db.Integer, db.ForeignKey('College.CODE')),
#                         db.Column('subject_course_code', db.Integer, db.ForeignKey('Course.Course_code'))
#                         )


# class Data(db.Model):
#     __tablename__ = 'College'
#     SR_NO = db.Column(db.Integer)
#     CODE = db.Column(db.Integer, primary_key=True)
#     INSTITUTE_NAME = db.Column(db.String(2500))
#     Following = db.relationship('Subject', secondary=data_subject, backref='followers')

#     def __repr__(self):
#         return f'<Data: {self.INSTITUTE_NAME}'


# @app.route('/map', methods=["POST"])
# def reduce():
#     if request.method=="POST":
#         file = request.files['file']
#         workbook = load_workbook(file)
#         sheet = workbook.active

#         for row in sheet.iter_rows(min_row=2, values_only=True):
#             excel_data = Data(SR_NO=row[0], CODE=row[1], INSTITUTE_NAME=str(row[2]))
#             db.session.add(excel_data)
#         db.session.commit()
#     return "hello"



# class Subject(db.Model):
#     __tablename__ = 'Course'
#     SR_no = db.Column(db.Integer)
#     Course_code = db.Column(db.Integer, primary_key=True)
#     Course_name = db.Column(db.String(1000))
    
#     def __repr__(self):
#         return f'<Subject: {self.Course_name}'


# @app.route('/come', methods=['POST'])
# def index():
#     if request.method == "POST":
#         file1 = request.files['size']
#         wb = load_workbook(file1)
#         sheet1 = wb.active

#         for row in sheet1.iter_rows(min_row=3, values_only=True):
#             excel_data1 = Subject(SR_no=row[0], Course_code=row[1], Course_name=str(row[2]))
#             db.session.add(excel_data1)
#         db.session.commit()
#     return {1:"Afsar Khan"}



class append(db.Model):
    SR_NO = db.Column(db.Integer)
    CODE = db.Column(db.Integer, primary_key=True)
    INSTITUTE_NAME = db.Column(db.String(150000))
    Course_code = db.Column(db.Integer, primary_key=True)
    Course_name = db.Column(db.String(100000))

@app.route('/tree', methods=["POST"])
def Leaf():
    if request.method == "POST":
        file2 = request.files['dear']
        wb2 = load_workbook(file2)
        sheet2 =wb2.active

        file3 = request.files['mumbai']
        wb3 = load_workbook(file3)
        sheet3 = wb3.active
        lst=[]
        for row in sheet2.iter_rows(min_row=2, values_only=True):
            for output in sheet3.iter_rows(min_row=2, values_only=True):
                print("$$$$$$$$$$$$$$$$$$$$$$$$iterated$$$$$$$$$$$$$$$$$$$$$$$$$")
                lst.append((row,output))
                print("!!!!!!!!!!!!!!!!!!!!!!!!!append!!!!!!!!!!!!!!!!!!!!!!!!!!")
                print("this si resut",lst)
                print("#########################lst##########################")
                excel_data3 = append(SR_NO=row[0], CODE=row[1], INSTITUTE_NAME=str(row[2]), Course_code=output[0], Course_name=str(output[1]))
                print("^^^^^^^^^^^^^^^^^^^^^^^^^excel_data3^^^^^^^^^^^^^^^^^^^^^^")
                db.session.add(excel_data3)
                print("------------------------------add-----------------------")
            db.session.commit()
            print("?????????????????????????commited???????????????????????????????????")
                
                # print(row)
                # print(output)
    return "Afsar Khan"

if __name__ == '__main__':
    app.run(debug=True)
    